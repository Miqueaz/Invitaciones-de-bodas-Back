from flask import Flask, request, jsonify
from flask_cors import CORS  # Importar Flask-CORS
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)
CORS(app)  # Habilitar CORS para todas las rutas

# Ruta al archivo Excel
EXCEL_FILE = 'asistencia.xlsx'

# Verificar si el archivo Excel existe, si no, crearlo
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    ws.append(['ID', 'Fecha', 'Hora'])  # Encabezados de la hoja
    wb.save(EXCEL_FILE)

@app.route('/registrar_asistencia', methods=['POST'])
def registrar_asistencia():
    try:
        # Verificar que la solicitud tenga JSON válido
        if not request.is_json:
            return jsonify({'error': 'La solicitud debe ser en formato JSON'}), 400

        # Obtener el ID del cuerpo de la solicitud
        data = request.json
        id_usuario = data.get('id')

        if not id_usuario:
            return jsonify({'error': 'ID no proporcionado'}), 400

        # Cargar el archivo Excel existente
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # Obtener la fecha y hora actual
        now = datetime.now()
        fecha = now.strftime('%Y-%m-%d')
        hora = now.strftime('%H:%M:%S')

        # Registrar la asistencia en el archivo Excel
        ws.append([id_usuario, fecha, hora])
        wb.save(EXCEL_FILE)

        return jsonify({'mensaje': 'Asistencia registrada correctamente'}), 200

    except Exception as e:
        # Manejar cualquier error inesperado
        return jsonify({'error': f'Error del servidor: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
